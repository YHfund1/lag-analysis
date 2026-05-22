from __future__ import annotations

import argparse
import json
import math
import shutil
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd


DEFAULT_SOURCE_FILE = Path("/Users/chenshuo/Nutstore Files/我的坚果云/中国_金融机构各项贷款余额_人民币_同比.xlsx")
DEFAULT_OUTPUT_DIR = Path("/Users/chenshuo/Desktop/结果")
DEFAULT_REPO_DIR = Path("/Users/chenshuo/Documents/Codex/2026-05-14/files-mentioned-by-the-user-xlsx/lag-analysis")
START_DATE = pd.Timestamp("2015-01-01")
EXCEL_NAME = "贷款指标_vs_10年国债收益率_相关性汇总.xlsx"
HTML_NAME = "贷款指标_vs_10年国债收益率_交互图.html"
RESULT_SCRIPT_NAME = "贷款指标_vs_10年国债收益率_分析脚本.py"

TARGET = "中债国债到期收益率:10年"
CANDIDATE_FACTORS = [
    "中国:金融机构各项贷款余额:人民币:同比",
    "信贷相较6个月前的变化",
    "信贷相较6个月前变化的同比",
    "信贷相较6个月前变化的同比增量",
    "信贷6个月脉冲",
]


def read_workbook(path: Path) -> pd.DataFrame:
    wb_values = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws_values = wb_values.active

    headers = [cell.value for cell in next(ws_values.iter_rows(min_row=2, max_row=2))]
    rows = []
    for row in ws_values.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        rows.append(row[: len(headers)])

    df = pd.DataFrame(rows, columns=headers)
    df = df.rename(columns={"指标名称": "日期"})
    df["日期"] = pd.to_datetime(df["日期"])
    df = df.sort_values("日期").reset_index(drop=True)
    for col in dict.fromkeys([TARGET, *CANDIDATE_FACTORS]):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def available_factors(df: pd.DataFrame) -> tuple[list[str], list[str]]:
    factors = [factor for factor in CANDIDATE_FACTORS if factor in df.columns]
    skipped = [factor for factor in CANDIDATE_FACTORS if factor not in df.columns]
    if TARGET not in df.columns:
        raise ValueError(f"源表缺少目标变量列：{TARGET}")
    if not factors:
        raise ValueError("源表没有可分析的候选因子列")
    return factors, skipped


def corr_for_lags(df: pd.DataFrame, factors: list[str]) -> tuple[pd.DataFrame, pd.DataFrame]:
    rows = []
    chart_rows = []

    for factor in factors:
        for lag in range(0, 21):
            shifted = df[factor].shift(lag)
            aligned = pd.DataFrame(
                {
                    "日期": df["日期"],
                    "target": df[TARGET],
                    "factor_shifted": shifted,
                    "factor_original_date": df["日期"].shift(lag),
                }
            )
            aligned = aligned[aligned["日期"] >= START_DATE].copy()
            pairs = aligned[["target", "factor_shifted"]].dropna()
            corr = pairs["target"].corr(pairs["factor_shifted"]) if len(pairs) >= 2 else math.nan
            rows.append(
                {
                    "因子": factor,
                    "因子领先10年国债收益率（月）": lag,
                    "相关系数": corr,
                    "相关系数绝对值": abs(corr) if pd.notna(corr) else math.nan,
                    "样本数": len(pairs),
                    "收益率样本开始": pairs.index.min() if len(pairs) else None,
                    "收益率样本结束": pairs.index.max() if len(pairs) else None,
                }
            )

            display = aligned.dropna(subset=["target", "factor_shifted"]).copy()
            for _, item in display.iterrows():
                chart_rows.append(
                    {
                        "factor": factor,
                        "lag": lag,
                        "date": item["日期"].strftime("%Y-%m-%d"),
                        "target": none_or_float(item["target"]),
                        "factorShifted": none_or_float(item["factor_shifted"]),
                        "factorOriginalDate": (
                            item["factor_original_date"].strftime("%Y-%m-%d")
                            if pd.notna(item["factor_original_date"])
                            else None
                        ),
                    }
                )

    corr_df = pd.DataFrame(rows)
    # Convert sample index back into real dates for the spreadsheet.
    if not corr_df.empty:
        corr_df["收益率样本开始"] = corr_df["收益率样本开始"].apply(
            lambda i: df.loc[i, "日期"].date().isoformat() if pd.notna(i) else ""
        )
        corr_df["收益率样本结束"] = corr_df["收益率样本结束"].apply(
            lambda i: df.loc[i, "日期"].date().isoformat() if pd.notna(i) else ""
        )
    return corr_df, pd.DataFrame(chart_rows)


def none_or_float(value):
    if value is None or pd.isna(value):
        return None
    return float(value)


def make_best_summary(corr_df: pd.DataFrame, factors: list[str]) -> pd.DataFrame:
    detailed = corr_df[corr_df["因子领先10年国债收益率（月）"].between(1, 20)].copy()
    best_rows = []
    for factor in factors:
        part = detailed[detailed["因子"] == factor].dropna(subset=["相关系数"])
        if part.empty:
            continue
        best_abs = part.loc[part["相关系数绝对值"].idxmax()]
        max_corr = part.loc[part["相关系数"].idxmax()]
        lowest_corr = part.loc[part["相关系数"].idxmin()]
        best_rows.append(
            {
                "因子": factor,
                "最强相关领先期（月，按绝对值）": int(best_abs["因子领先10年国债收益率（月）"]),
                "最强相关系数": best_abs["相关系数"],
                "最强相关方向": "正相关" if best_abs["相关系数"] >= 0 else "负相关",
                "样本数": int(best_abs["样本数"]),
                "样本区间": f'{best_abs["收益率样本开始"]} 至 {best_abs["收益率样本结束"]}',
                "最大相关领先期（月）": int(max_corr["因子领先10年国债收益率（月）"]),
                "最大相关系数": max_corr["相关系数"],
                "最低相关领先期（月）": int(lowest_corr["因子领先10年国债收益率（月）"]),
                "最低相关系数": lowest_corr["相关系数"],
            }
        )
    return pd.DataFrame(best_rows)


def make_data_sheet(df: pd.DataFrame, factors: list[str]) -> pd.DataFrame:
    cols = ["日期", TARGET, *factors]
    out = df.loc[df["日期"] >= START_DATE, cols].copy()
    out["日期"] = out["日期"].dt.date.astype(str)
    return out


def write_excel(
    best_df: pd.DataFrame,
    corr_df: pd.DataFrame,
    data_df: pd.DataFrame,
    factors: list[str],
    skipped_factors: list[str],
    source_file: Path,
    output: Path,
) -> None:
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        readme = pd.DataFrame(
            [
                ["源文件", str(source_file)],
                ["分析开始时间", START_DATE.date().isoformat()],
                ["目标变量", TARGET],
                ["滞后定义", "因子领先10年国债收益率x个月：用t-x期因子对应t期10年国债收益率"],
                ["汇总口径", "最强相关领先期按相关系数绝对值最大判断；同时给出最大相关系数和最低相关系数"],
                ["已分析因子", "；".join(factors)],
                ["源表缺失并跳过的候选因子", "；".join(skipped_factors) if skipped_factors else "无"],
            ],
            columns=["项目", "说明"],
        )
        readme.to_excel(writer, sheet_name="说明", index=False)
        best_df.to_excel(writer, sheet_name="最优领先期汇总", index=False)
        corr_df.to_excel(writer, sheet_name="0-20期相关性明细", index=False)
        data_df.to_excel(writer, sheet_name="2015以来原始数据", index=False)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            for col in ws.columns:
                max_len = 0
                letter = col[0].column_letter
                for cell in col:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)


def build_html(
    chart_df: pd.DataFrame,
    corr_df: pd.DataFrame,
    best_df: pd.DataFrame,
    factors: list[str],
    output: Path,
) -> None:
    corr_payload = {}
    for factor in factors:
        corr_payload[factor] = {}
        part = corr_df[corr_df["因子"] == factor]
        for _, row in part.iterrows():
            corr_payload[factor][int(row["因子领先10年国债收益率（月）"])] = {
                "corr": none_or_float(row["相关系数"]),
                "n": int(row["样本数"]),
                "start": row["收益率样本开始"],
                "end": row["收益率样本结束"],
            }

    series_payload = {}
    for factor in factors:
        series_payload[factor] = {}
        for lag in range(0, 21):
            part = chart_df[(chart_df["factor"] == factor) & (chart_df["lag"] == lag)]
            series_payload[factor][lag] = part[
                ["date", "target", "factorShifted", "factorOriginalDate"]
            ].to_dict(orient="records")

    best_payload = best_df.to_dict(orient="records")

    data_json = json.dumps(
        {
            "target": TARGET,
            "factors": factors,
            "corr": corr_payload,
            "series": series_payload,
            "best": best_payload,
            "startDate": START_DATE.date().isoformat(),
        },
        ensure_ascii=False,
        allow_nan=False,
    )

    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>贷款指标与10年国债收益率相关性分析</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --panel: #ffffff;
      --ink: #1f2933;
      --muted: #667085;
      --grid: #d8dee8;
      --blue: #2563eb;
      --green: #0f9f6e;
      --border: #d0d7e2;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
      color: var(--ink);
      background: var(--bg);
    }}
    header {{
      padding: 24px 28px 12px;
      background: var(--panel);
      border-bottom: 1px solid var(--border);
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 24px;
      line-height: 1.25;
      letter-spacing: 0;
    }}
    .subtitle {{
      color: var(--muted);
      font-size: 14px;
      line-height: 1.7;
    }}
    main {{
      padding: 20px 28px 28px;
      max-width: 1280px;
      margin: 0 auto;
    }}
    .summary {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .metric {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 14px;
      min-height: 108px;
    }}
    .metric .name {{
      font-size: 13px;
      color: var(--muted);
      margin-bottom: 8px;
      line-height: 1.45;
      overflow-wrap: anywhere;
    }}
    .metric .value {{
      font-size: 24px;
      font-weight: 700;
      margin-bottom: 4px;
    }}
    .metric .detail {{
      font-size: 13px;
      color: var(--muted);
      line-height: 1.45;
    }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--border);
      border-radius: 8px;
      padding: 16px;
      margin-bottom: 18px;
    }}
    .controls {{
      display: grid;
      grid-template-columns: minmax(280px, 1fr) minmax(280px, 1fr);
      gap: 16px;
      align-items: end;
      margin-bottom: 12px;
    }}
    label {{
      display: block;
      font-size: 13px;
      color: var(--muted);
      margin-bottom: 6px;
    }}
    select, input[type="range"] {{
      width: 100%;
    }}
    select {{
      height: 38px;
      border: 1px solid var(--border);
      border-radius: 6px;
      background: white;
      color: var(--ink);
      font-size: 14px;
      padding: 0 10px;
    }}
    .lagLine {{
      display: flex;
      gap: 12px;
      align-items: center;
    }}
    .lagValue {{
      min-width: 74px;
      text-align: right;
      font-weight: 700;
      color: var(--ink);
    }}
    .stats {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin: 10px 0 12px;
      color: var(--muted);
      font-size: 13px;
    }}
    .pill {{
      border: 1px solid var(--border);
      border-radius: 999px;
      padding: 6px 10px;
      background: #fbfcfe;
    }}
    .chartActions {{
      display: flex;
      justify-content: flex-end;
      margin: 2px 0 12px;
    }}
    .saveButton {{
      height: 34px;
      border: 1px solid var(--border);
      border-radius: 6px;
      background: #ffffff;
      color: var(--ink);
      cursor: pointer;
      font-size: 13px;
      font-weight: 600;
      padding: 0 12px;
    }}
    .saveButton:hover {{
      border-color: #98a2b3;
      background: #f8fafc;
    }}
    .saveButton:active {{
      transform: translateY(1px);
    }}
    .chartWrap {{
      position: relative;
      width: 100%;
      height: 560px;
      border: 1px solid var(--border);
      border-radius: 8px;
      overflow: hidden;
      background: white;
    }}
    svg {{
      width: 100%;
      height: 100%;
      display: block;
    }}
    .tooltip {{
      position: absolute;
      pointer-events: none;
      display: none;
      background: rgba(17, 24, 39, 0.94);
      color: white;
      padding: 8px 10px;
      border-radius: 6px;
      font-size: 12px;
      line-height: 1.55;
      max-width: 320px;
      z-index: 10;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
      background: white;
    }}
    th, td {{
      border-bottom: 1px solid var(--border);
      padding: 9px 10px;
      text-align: right;
      white-space: nowrap;
    }}
    th:first-child, td:first-child, th:nth-child(2), td:nth-child(2) {{
      text-align: left;
    }}
    th {{
      color: var(--muted);
      font-weight: 600;
      background: #fbfcfe;
    }}
    .legend {{
      display: flex;
      gap: 16px;
      align-items: center;
      margin: 10px 0 0;
      color: var(--muted);
      font-size: 13px;
    }}
    .legend span::before {{
      content: "";
      display: inline-block;
      width: 18px;
      height: 3px;
      border-radius: 999px;
      margin-right: 6px;
      vertical-align: middle;
      background: var(--blue);
    }}
    .legend span:last-child::before {{
      background: var(--green);
    }}
    @media (max-width: 860px) {{
      header {{ padding: 20px 16px 10px; }}
      main {{ padding: 16px; }}
      .summary {{ grid-template-columns: 1fr; }}
      .controls {{ grid-template-columns: 1fr; }}
      .chartWrap {{ height: 460px; }}
      table {{ font-size: 12px; }}
      th, td {{ padding: 8px 6px; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>贷款指标与10年国债收益率相关性分析</h1>
    <div class="subtitle">样本从 2015-01-01 开始。滞后定义：把因子向后移动 x 期，即用 t-x 期因子对应 t 期 10 年国债收益率；相关性为 Pearson 相关系数。</div>
  </header>
  <main>
    <section class="summary" id="summary"></section>

    <section class="panel">
      <div class="controls">
        <div>
          <label for="factorSelect">指标</label>
          <select id="factorSelect"></select>
        </div>
        <div>
          <label for="lagRange">因子领先期</label>
          <div class="lagLine">
            <input id="lagRange" type="range" min="0" max="20" step="1" value="0">
            <div class="lagValue"><span id="lagValue">0</span> 期</div>
          </div>
        </div>
      </div>
      <div class="stats">
        <div class="pill">相关系数：<strong id="corrValue">-</strong></div>
        <div class="pill">样本数：<strong id="sampleValue">-</strong></div>
        <div class="pill">收益率样本：<strong id="rangeValue">-</strong></div>
      </div>
      <div class="chartActions">
        <button class="saveButton" id="saveChartButton" type="button" title="保存当前折线图为PNG图片">保存图片</button>
      </div>
      <div class="chartWrap" id="chartWrap">
        <svg id="chart" role="img" aria-label="指标与10年国债收益率折线图"></svg>
        <div class="tooltip" id="tooltip"></div>
      </div>
      <div class="legend">
        <span>10年国债收益率（左轴，%）</span>
        <span>滞后后因子（右轴，原单位）</span>
      </div>
    </section>

    <section class="panel">
      <table>
        <thead>
          <tr>
            <th>因子</th>
            <th>最强相关领先期</th>
            <th>相关系数</th>
            <th>方向</th>
            <th>样本数</th>
            <th>最大相关</th>
            <th>最低相关</th>
          </tr>
        </thead>
        <tbody id="bestTable"></tbody>
      </table>
    </section>
  </main>

  <script>
    const payload = {data_json};
    const factorSelect = document.getElementById('factorSelect');
    const lagRange = document.getElementById('lagRange');
    const lagValue = document.getElementById('lagValue');
    const corrValue = document.getElementById('corrValue');
    const sampleValue = document.getElementById('sampleValue');
    const rangeValue = document.getElementById('rangeValue');
    const svg = document.getElementById('chart');
    const chartWrap = document.getElementById('chartWrap');
    const tooltip = document.getElementById('tooltip');
    const saveChartButton = document.getElementById('saveChartButton');

    const fmt = new Intl.NumberFormat('zh-CN', {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});
    const fmt2 = new Intl.NumberFormat('zh-CN', {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});
    const dateFmt = new Intl.DateTimeFormat('zh-CN', {{ year: 'numeric', month: '2-digit' }});

    function formatCorr(value) {{
      if (value === null || Number.isNaN(value)) return '-';
      return value.toFixed(2);
    }}

    function init() {{
      payload.factors.forEach((factor) => {{
        const option = document.createElement('option');
        option.value = factor;
        option.textContent = factor;
        factorSelect.appendChild(option);
      }});
      renderSummary();
      renderBestTable();
      update();
    }}

    function renderSummary() {{
      const root = document.getElementById('summary');
      root.innerHTML = '';
      payload.best.forEach((row) => {{
        const card = document.createElement('div');
        card.className = 'metric';
        card.innerHTML = `
          <div class="name">${{escapeHtml(row['因子'])}}</div>
          <div class="value">${{row['最强相关领先期（月，按绝对值）']}} 期</div>
          <div class="detail">r=${{formatCorr(row['最强相关系数'])}}，${{escapeHtml(row['最强相关方向'])}}，n=${{row['样本数']}}</div>
        `;
        root.appendChild(card);
      }});
    }}

    function renderBestTable() {{
      const tbody = document.getElementById('bestTable');
      tbody.innerHTML = payload.best.map((row) => `
        <tr>
          <td>${{escapeHtml(row['因子'])}}</td>
          <td>${{row['最强相关领先期（月，按绝对值）']}} 期</td>
          <td>${{formatCorr(row['最强相关系数'])}}</td>
          <td>${{escapeHtml(row['最强相关方向'])}}</td>
          <td>${{row['样本数']}}</td>
          <td>${{row['最大相关领先期（月）']}} 期 / ${{formatCorr(row['最大相关系数'])}}</td>
          <td>${{row['最低相关领先期（月）']}} 期 / ${{formatCorr(row['最低相关系数'])}}</td>
        </tr>
      `).join('');
    }}

    function update() {{
      const factor = factorSelect.value;
      const lag = Number(lagRange.value);
      lagValue.textContent = lag;
      const stat = payload.corr[factor][lag];
      corrValue.textContent = formatCorr(stat.corr);
      sampleValue.textContent = stat.n;
      rangeValue.textContent = `${{stat.start}} 至 ${{stat.end}}`;
      drawChart(payload.series[factor][lag], factor, lag);
    }}

    function drawChart(data, factorLabel, lag) {{
      const rect = chartWrap.getBoundingClientRect();
      const width = Math.max(360, rect.width);
      const height = Math.max(320, rect.height);
      svg.setAttribute('viewBox', `0 0 ${{width}} ${{height}}`);
      svg.innerHTML = '';

      const margin = {{ top: 36, right: 100, bottom: 62, left: 78 }};
      const plotW = width - margin.left - margin.right;
      const plotH = height - margin.top - margin.bottom;
      if (!data || data.length < 2) return;

      const dates = data.map(d => new Date(d.date + 'T00:00:00'));
      const xs = dates.map(d => d.getTime());
      const y1 = data.map(d => d.target);
      const y2 = data.map(d => d.factorShifted);

      const xMin = Math.min(...xs), xMax = Math.max(...xs);
      const y1Domain = paddedDomain(y1);
      const y2Domain = paddedDomain(y2);

      const xScale = (x) => margin.left + ((x - xMin) / (xMax - xMin || 1)) * plotW;
      const y1Scale = (y) => margin.top + (1 - (y - y1Domain[0]) / (y1Domain[1] - y1Domain[0] || 1)) * plotH;
      const y2Scale = (y) => margin.top + (1 - (y - y2Domain[0]) / (y2Domain[1] - y2Domain[0] || 1)) * plotH;

      drawGrid(width, height, margin, plotW, plotH, y1Domain, y1Scale, y2Domain, y2Scale, xMin, xMax, xScale);
      drawPath(data, (d) => xScale(new Date(d.date + 'T00:00:00').getTime()), (d) => y1Scale(d.target), '#2563eb', 2.2);
      drawPath(data, (d) => xScale(new Date(d.date + 'T00:00:00').getTime()), (d) => y2Scale(d.factorShifted), '#0f9f6e', 2.2);

      addText(margin.left, 24, '10年国债收益率 %', '#2563eb', 'start', 16, 800);
      addText(width - margin.right, 24, `${{factorLabel}}，领先 ${{lag}} 期`, '#0f9f6e', 'end', 16, 800);

      const overlay = document.createElementNS('http://www.w3.org/2000/svg', 'rect');
      overlay.setAttribute('x', margin.left);
      overlay.setAttribute('y', margin.top);
      overlay.setAttribute('width', plotW);
      overlay.setAttribute('height', plotH);
      overlay.setAttribute('fill', 'transparent');
      overlay.addEventListener('mousemove', (event) => {{
        const point = svg.createSVGPoint();
        point.x = event.clientX;
        point.y = event.clientY;
        const cursor = point.matrixTransform(svg.getScreenCTM().inverse());
        const ratio = Math.min(1, Math.max(0, (cursor.x - margin.left) / plotW));
        const targetTime = xMin + ratio * (xMax - xMin);
        let bestIndex = 0;
        let bestDistance = Infinity;
        xs.forEach((x, i) => {{
          const distance = Math.abs(x - targetTime);
          if (distance < bestDistance) {{
            bestDistance = distance;
            bestIndex = i;
          }}
        }});
        const d = data[bestIndex];
        const x = xScale(xs[bestIndex]);
        const yTarget = y1Scale(d.target);
        drawHover(x, yTarget, margin.top, margin.top + plotH);
        tooltip.style.display = 'block';
        const maxLeft = Math.max(8, rect.width - 320);
        tooltip.style.left = `${{Math.min(maxLeft, Math.max(8, event.clientX - rect.left + 14))}}px`;
        tooltip.style.top = `${{Math.max(8, event.clientY - rect.top + 14)}}px`;
        tooltip.innerHTML = `
          <strong>${{escapeHtml(d.date)}}</strong><br>
          10年国债收益率：${{fmt.format(d.target)}}%<br>
          ${{escapeHtml(factorLabel)}}：${{fmt2.format(d.factorShifted)}}<br>
          因子原始日期：${{escapeHtml(d.factorOriginalDate || '-')}}
        `;
      }});
      overlay.addEventListener('mouseleave', () => {{
        tooltip.style.display = 'none';
        svg.querySelectorAll('.hover').forEach(el => el.remove());
      }});
      svg.appendChild(overlay);
    }}

    function saveChartImage() {{
      const viewBox = (svg.getAttribute('viewBox') || '').split(/\\s+/).map(Number);
      if (viewBox.length !== 4 || viewBox.some(Number.isNaN)) return;

      const width = viewBox[2];
      const height = viewBox[3];
      const clonedSvg = svg.cloneNode(true);
      clonedSvg.querySelectorAll('.hover').forEach(el => el.remove());
      clonedSvg.querySelectorAll('rect[fill="transparent"]').forEach(el => el.remove());
      clonedSvg.setAttribute('xmlns', 'http://www.w3.org/2000/svg');
      clonedSvg.setAttribute('width', width);
      clonedSvg.setAttribute('height', height);

      const background = document.createElementNS('http://www.w3.org/2000/svg', 'rect');
      background.setAttribute('x', 0);
      background.setAttribute('y', 0);
      background.setAttribute('width', width);
      background.setAttribute('height', height);
      background.setAttribute('fill', '#ffffff');
      clonedSvg.insertBefore(background, clonedSvg.firstChild);

      const style = document.createElementNS('http://www.w3.org/2000/svg', 'style');
      style.textContent = 'text {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif; }}';
      clonedSvg.insertBefore(style, clonedSvg.firstChild);

      const svgText = new XMLSerializer().serializeToString(clonedSvg);
      const svgBlob = new Blob([svgText], {{ type: 'image/svg+xml;charset=utf-8' }});
      const svgUrl = URL.createObjectURL(svgBlob);
      const image = new Image();
      image.onload = () => {{
        const scale = Math.max(2, Math.min(3, window.devicePixelRatio || 2));
        const canvas = document.createElement('canvas');
        canvas.width = Math.round(width * scale);
        canvas.height = Math.round(height * scale);
        const context = canvas.getContext('2d');
        context.fillStyle = '#ffffff';
        context.fillRect(0, 0, canvas.width, canvas.height);
        context.drawImage(image, 0, 0, canvas.width, canvas.height);
        URL.revokeObjectURL(svgUrl);

        canvas.toBlob((blob) => {{
          if (!blob) return;
          const factor = factorSelect.value;
          const lag = Number(lagRange.value);
          const fileUrl = URL.createObjectURL(blob);
          const link = document.createElement('a');
          link.href = fileUrl;
          link.download = `lag-analysis_${{safeFilename(factor)}}_lead-${{lag}}.png`;
          document.body.appendChild(link);
          link.click();
          link.remove();
          setTimeout(() => URL.revokeObjectURL(fileUrl), 1000);
        }}, 'image/png');
      }};
      image.onerror = () => {{
        URL.revokeObjectURL(svgUrl);
        alert('图片生成失败，请刷新页面后重试。');
      }};
      image.src = svgUrl;
    }}

    function drawGrid(width, height, margin, plotW, plotH, y1Domain, y1Scale, y2Domain, y2Scale, xMin, xMax, xScale) {{
      const gridGroup = document.createElementNS('http://www.w3.org/2000/svg', 'g');
      svg.appendChild(gridGroup);
      for (let i = 0; i <= 5; i++) {{
        const yValue = y1Domain[0] + (i / 5) * (y1Domain[1] - y1Domain[0]);
        const y = y1Scale(yValue);
        line(margin.left, y, margin.left + plotW, y, '#d8dee8', 1, gridGroup);
        addText(margin.left - 12, y + 5, fmt.format(yValue), '#111827', 'end', 15, 800);
        const rightValue = y2Domain[0] + (i / 5) * (y2Domain[1] - y2Domain[0]);
        addText(width - margin.right + 12, y + 5, compactNumber(rightValue), '#111827', 'start', 15, 800);
      }}
      for (let i = 0; i <= 6; i++) {{
        const xValue = xMin + (i / 6) * (xMax - xMin);
        const x = xScale(xValue);
        line(x, margin.top, x, margin.top + plotH, '#eef1f5', 1, gridGroup);
        addText(x, height - 24, dateFmt.format(new Date(xValue)), '#111827', 'middle', 15, 800);
      }}
      line(margin.left, margin.top, margin.left, margin.top + plotH, '#9aa4b2', 1, gridGroup);
      line(width - margin.right, margin.top, width - margin.right, margin.top + plotH, '#9aa4b2', 1, gridGroup);
      line(margin.left, margin.top + plotH, margin.left + plotW, margin.top + plotH, '#9aa4b2', 1, gridGroup);
    }}

    function drawPath(data, xFn, yFn, color, strokeWidth) {{
      const path = document.createElementNS('http://www.w3.org/2000/svg', 'path');
      const d = data.map((row, i) => `${{i === 0 ? 'M' : 'L'}}${{xFn(row).toFixed(2)}},${{yFn(row).toFixed(2)}}`).join(' ');
      path.setAttribute('d', d);
      path.setAttribute('fill', 'none');
      path.setAttribute('stroke', color);
      path.setAttribute('stroke-width', strokeWidth);
      path.setAttribute('stroke-linejoin', 'round');
      path.setAttribute('stroke-linecap', 'round');
      svg.appendChild(path);
    }}

    function drawHover(x, y, yTop, yBottom) {{
      svg.querySelectorAll('.hover').forEach(el => el.remove());
      const group = document.createElementNS('http://www.w3.org/2000/svg', 'g');
      group.setAttribute('class', 'hover');
      line(x, yTop, x, yBottom, '#111827', 1, group, '4 4');
      const dot = document.createElementNS('http://www.w3.org/2000/svg', 'circle');
      dot.setAttribute('cx', x);
      dot.setAttribute('cy', y);
      dot.setAttribute('r', 4);
      dot.setAttribute('fill', '#2563eb');
      group.appendChild(dot);
      svg.appendChild(group);
    }}

    function line(x1, y1, x2, y2, color, width, parent, dash = '') {{
      const el = document.createElementNS('http://www.w3.org/2000/svg', 'line');
      el.setAttribute('x1', x1);
      el.setAttribute('y1', y1);
      el.setAttribute('x2', x2);
      el.setAttribute('y2', y2);
      el.setAttribute('stroke', color);
      el.setAttribute('stroke-width', width);
      if (dash) el.setAttribute('stroke-dasharray', dash);
      (parent || svg).appendChild(el);
      return el;
    }}

    function addText(x, y, text, color, anchor, size, weight) {{
      const el = document.createElementNS('http://www.w3.org/2000/svg', 'text');
      el.setAttribute('x', x);
      el.setAttribute('y', y);
      el.setAttribute('fill', color);
      el.setAttribute('text-anchor', anchor);
      el.setAttribute('font-size', size);
      el.setAttribute('font-weight', weight);
      el.textContent = text;
      svg.appendChild(el);
      return el;
    }}

    function paddedDomain(values) {{
      const clean = values.filter(v => v !== null && !Number.isNaN(v));
      let min = Math.min(...clean), max = Math.max(...clean);
      if (min === max) {{
        min -= 1;
        max += 1;
      }}
      const pad = (max - min) * 0.08;
      return [min - pad, max + pad];
    }}

    function compactNumber(value) {{
      const abs = Math.abs(value);
      if (abs >= 10000) return `${{fmt.format(value / 10000)}}万`;
      return fmt.format(value);
    }}

    function escapeHtml(value) {{
      return String(value)
        .replaceAll('&', '&amp;')
        .replaceAll('<', '&lt;')
        .replaceAll('>', '&gt;')
        .replaceAll('"', '&quot;')
        .replaceAll("'", '&#039;');
    }}

    function safeFilename(value) {{
      return String(value)
        .replace(/[\\\\/:*?"<>|\\s]+/g, '_')
        .replace(/_+/g, '_')
        .replace(/^_|_$/g, '')
        .slice(0, 80);
    }}

    factorSelect.addEventListener('change', update);
    lagRange.addEventListener('input', update);
    saveChartButton.addEventListener('click', saveChartImage);
    window.addEventListener('resize', update);
    init();
  </script>
</body>
</html>
"""
    output.write_text(html, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate lag-correlation outputs from the source workbook and publish them to GitHub Pages."
    )
    parser.add_argument("--source-file", type=Path, default=DEFAULT_SOURCE_FILE)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--repo-dir", type=Path, default=DEFAULT_REPO_DIR)
    parser.add_argument("--start-date", default=START_DATE.date().isoformat())
    parser.add_argument("--commit-message", default="Update lag analysis results")
    parser.add_argument("--no-git", action="store_true", help="Generate outputs and copy repository files, but skip commit/push.")
    parser.add_argument("--dry-run", action="store_true", help="Generate outputs, then show git changes without committing or pushing.")
    return parser.parse_args()


def run_command(cmd: list[str], cwd: Path | None = None) -> subprocess.CompletedProcess[str]:
    result = subprocess.run(cmd, cwd=cwd, text=True, capture_output=True)
    if result.returncode != 0:
        command = " ".join(cmd)
        raise RuntimeError(f"Command failed: {command}\nSTDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}")
    return result


def sync_repository_files(script_file: Path, html_output: Path, excel_output: Path, repo_dir: Path) -> None:
    if not repo_dir.exists():
        raise FileNotFoundError(f"Git repository directory does not exist: {repo_dir}")
    shutil.copy2(html_output, repo_dir / "index.html")
    shutil.copy2(excel_output, repo_dir / EXCEL_NAME)
    shutil.copy2(script_file, repo_dir / "analyze_credit_yield.py")


def git_status(repo_dir: Path) -> str:
    return run_command(["git", "status", "--short"], cwd=repo_dir).stdout.strip()


def publish_repository(repo_dir: Path, commit_message: str, dry_run: bool) -> None:
    add_paths = ["index.html", "analyze_credit_yield.py", EXCEL_NAME]
    add_paths.extend(path.name for path in [repo_dir / "README.md", repo_dir / ".nojekyll"] if path.exists())
    run_command(["git", "add", *add_paths], cwd=repo_dir)
    status = git_status(repo_dir)
    if not status:
        print("git_status=clean")
        return
    print("git_status_before_commit:\n" + status)
    if dry_run:
        print("dry_run=true; skipped commit and push")
        return
    run_command(["git", "commit", "-m", commit_message], cwd=repo_dir)
    run_command(["git", "push", "origin", "main"], cwd=repo_dir)
    run_command(["git", "push", "origin", "main:gh-pages"], cwd=repo_dir)
    remote = run_command(["git", "ls-remote", "origin", "refs/heads/main", "refs/heads/gh-pages"], cwd=repo_dir)
    print("remote_refs:\n" + remote.stdout.strip())


def main() -> None:
    global START_DATE

    args = parse_args()
    START_DATE = pd.Timestamp(args.start_date)
    source_file = args.source_file.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    repo_dir = args.repo_dir.expanduser().resolve()
    script_file = Path(__file__).resolve()

    output_dir.mkdir(parents=True, exist_ok=True)
    df = read_workbook(source_file)
    factors, skipped_factors = available_factors(df)
    corr_df, chart_df = corr_for_lags(df, factors)
    best_df = make_best_summary(corr_df, factors)
    data_df = make_data_sheet(df, factors)

    excel_output = output_dir / EXCEL_NAME
    html_output = output_dir / HTML_NAME
    result_script_output = output_dir / RESULT_SCRIPT_NAME

    write_excel(best_df, corr_df, data_df, factors, skipped_factors, source_file, excel_output)
    build_html(chart_df, corr_df, best_df, factors, html_output)
    shutil.copy2(script_file, result_script_output)
    sync_repository_files(script_file, html_output, excel_output, repo_dir)

    print(f"excel={excel_output}")
    print(f"html={html_output}")
    print(f"script={result_script_output}")
    print(f"repo={repo_dir}")
    print("analyzed_factors=" + " | ".join(factors))
    if skipped_factors:
        print("skipped_missing_factors=" + " | ".join(skipped_factors))
    print(best_df.to_string(index=False))
    if args.no_git:
        print("no_git=true; skipped commit and push")
    else:
        publish_repository(repo_dir, args.commit_message, args.dry_run)


if __name__ == "__main__":
    main()
